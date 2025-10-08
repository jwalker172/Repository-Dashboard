from flask import Flask, render_template, request, jsonify
import pandas as pd
from openpyxl import load_workbook
import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

app = Flask(__name__)

EXCEL_FILE = 'data.xlsx'
SHEET_NAMES = ['BP-P', 'FR-FZ', 'GYL-ME', 'LBTF', 'PFC', 'PSC']


@app.route('/')
def index():
    return render_template('index.html', sheet_names=SHEET_NAMES)

@app.route('/get_pe_re_list', methods=['POST'])
def get_pe_re_list():
    data = request.json
    sheet = data.get('sheet')
    if sheet not in SHEET_NAMES:
        return jsonify({'error': 'Invalid sheet name'}), 400

    df = pd.read_excel(EXCEL_FILE, sheet_name=sheet, header=1)

    if 'PE/RE' not in df.columns:
        return jsonify({'error': 'Missing PE/RE column'}), 500

    pe_re_list = sorted(df['PE/RE'].dropna().unique())
    return jsonify(pe_re_list)
    
    

@app.route('/get_total_gain', methods=['POST'])
def get_total_gain():
    data = request.json
    sheet = data.get('sheet')
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb[sheet]

        total = 0
        for i in range(3, 300):  # I3 to I29
            cell = ws[f'I{i}']
            value = cell.value

            # Try to coerce strings to float if needed
            try:
                numeric_value = float(value)
            except (TypeError, ValueError):
                numeric_value = 0  # Treat non-numeric as 0

            # Optionally: convert the cell to a real number (fixing Excel recognition)
            cell.value = numeric_value

            total += numeric_value

        # Set the cleaned-up total in I1
        ws['I1'] = total

        wb.save(EXCEL_FILE)
        return jsonify({'total_gain': total})
    except Exception as e:
        return jsonify({'error': str(e)})




@app.route('/get_wells', methods=['POST'])
def get_wells():
    data = request.json
    sheet = data.get('sheet')
    pe_re = data.get('pe_re')

    try:
        xls = pd.ExcelFile(EXCEL_FILE)
        if sheet not in xls.sheet_names:
            return jsonify({'error': 'Invalid sheet name'}), 400

        df = pd.read_excel(xls, sheet_name=sheet, header=1)
        df.columns = df.columns.str.strip()

        if 'PE/RE' not in df.columns:
            return jsonify({'error': 'PE/RE column missing'}), 500

        df['PE/RE'] = df['PE/RE'].astype(str).str.strip().str.lower()
        pe_re_clean = pe_re.strip().lower()

        filtered = df[df['PE/RE'] == pe_re_clean]

        # Convert datetime columns to strings, avoiding NaTType issues
        for col in filtered.select_dtypes(include=['datetime']):
            filtered[col] = filtered[col].dt.strftime('%Y-%m-%d')

        filtered = filtered.fillna('')  # Fill other NaNs
        wells = filtered.to_dict(orient='records')

        if not wells:
            return jsonify({'error': f'No wells found for PE/RE: {pe_re} in sheet: {sheet}'}), 404

        return jsonify(wells)

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/save_well', methods=['POST'])
def save_well():
    data = request.json
    sheet_name = data.get('sheet')
    well = data.get('Well')

    if sheet_name not in SHEET_NAMES:
        return jsonify({'error': 'Invalid sheet name'}), 400
    if not well:
        return jsonify({'error': 'No well data provided'}), 400

    wb = load_workbook(EXCEL_FILE)
    ws = wb[sheet_name]
    
    # Get column headers
    headers = [cell.value for cell in ws[2]]
    print("Headers:", headers)
    try:
        well_name_col_index = headers.index('Well') + 1
    except ValueError:
        return jsonify({'error': "'Well' column not found in headers"}), 400


    # Find row with matching 'Well' name
    target_row = None
    for row in ws.iter_rows(min_row=2, max_col=len(headers), values_only=False):
        if str(row[well_name_col_index - 1].value).strip() == str(well.get('Well')).strip():
            target_row = row
            break

    # If not found, append a new row
    if not target_row:
        new_row_idx = ws.max_row + 1
        for col_index, header in enumerate(headers, start=1):
            ws.cell(row=new_row_idx, column=col_index).value = well.get(header, "")
    else:
        for col_index, header in enumerate(headers, start=1):
            target_row[col_index - 1].value = well.get(header, "")

    wb.save(EXCEL_FILE)
    return jsonify({'success': True})


@app.route('/save_well2', methods=['POST'])
def save_well2():
    data = request.json
    sheet_name = data.get('sheet')
    well = data.get('Well')

    if sheet_name not in SHEET_NAMES:
        return jsonify({'error': 'Invalid sheet name'}), 400
    if not well:
        return jsonify({'error': 'No well data provided'}), 400

    wb = load_workbook(EXCEL_FILE)
    ws = wb[sheet_name]
    
    # Get column headers
    headers = [cell.value for cell in ws[2]]
    print("Headers:", headers)
    try:
        well_name_col_index = headers.index('Well') + 1
    except ValueError:
        return jsonify({'error': "'Well' column not found in headers"}), 400


    # Find row with matching 'Well' name
    target_row = None
    for row in ws.iter_rows(min_row=2, max_col=len(headers), values_only=False):
        if str(row[well_name_col_index - 1].value).strip() == str(well.get('Well')).strip():
            target_row = row
            break

    # If not found, append a new row
    if not target_row:
        new_row_idx = ws.max_row + 1
        for col_index, header in enumerate(headers, start=1):
            ws.cell(row=new_row_idx, column=col_index).value = well.get(header, "")
    else:
        for col_index, header in enumerate(headers, start=1):
            target_row[col_index - 1].value = well.get(header, "")

    wb.save(EXCEL_FILE)
    return jsonify({'success': True})


@app.route('/add_well', methods=['POST'])
def add_well():
    data = request.json
    sheet = data.get('sheet')
    new_well = data.get('well')

    if sheet not in SHEET_NAMES:
        return jsonify({'error': 'Invalid sheet name'}), 400
    if not new_well:
        return jsonify({'error': 'No new well data provided'}), 400

    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb[sheet]

        # Get headers from row 2
        headers = [cell.value for cell in ws[2]]
        print("Headers:", headers)

        # Find the first empty row (where column 1 is empty)
        first_empty_row = None
        for row in range(3, ws.max_row + 2):  # Start from row 3 (assuming headers are in row 2)
            if ws.cell(row=row, column=1).value in [None, ""]:
                first_empty_row = row
                break

        if first_empty_row is None:
            return jsonify({'error': 'No empty row found'}), 500

        # Write the new well data
        for col_index, header in enumerate(headers, start=1):
            ws.cell(row=first_empty_row, column=col_index).value = new_well.get(header, "")

        wb.save(EXCEL_FILE)
        return jsonify({'success': True})

    except Exception as e:
        return jsonify({'error': str(e)}), 500



@app.route('/delete_well', methods=['POST'])
def delete_well():
    data = request.json
    sheet_name = data.get('sheet')
    well_name = data.get('well_name')

    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb[sheet_name]

        for row in ws.iter_rows(min_row=2):
            cell_value = row[3].value  # ✅ Column D (0-indexed as 3)
            if cell_value is None:
                continue  # Skip rows with blank well names

            if str(cell_value).strip() == well_name.strip():
                ws.delete_rows(row[0].row)
                wb.save(EXCEL_FILE)
                return jsonify({'success': True})

        return jsonify({'success': False, 'error': 'Well not found'})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/move_to_delete', methods=['POST'])
def move_to_delete():
    data = request.json
    sheet_name = data.get('sheet')
    well_name = data.get('well_name')

    if sheet_name not in SHEET_NAMES:
        return jsonify({'error': 'Invalid sheet name'}), 400

    try:
        wb = load_workbook(EXCEL_FILE)
        source_ws = wb[sheet_name]
        deleted_ws = wb['DELETED']

        # Get headers from row 2
        headers = [cell.value for cell in source_ws[2]]

        # Find the row with the matching well name (in column D, index 3)
        source_row_data = None
        for row in source_ws.iter_rows(min_row=3, values_only=False):
            cell_value = row[3].value  # Column D
            if cell_value and str(cell_value).strip() == well_name.strip():
                source_row_data = [cell.value for cell in row]
                source_ws.delete_rows(row[0].row)
                break

        if not source_row_data:
            return jsonify({'error': 'Well not found'}), 404

        # Find the first empty row in DELETED sheet (starting from row 3)
        for row_idx in range(3, deleted_ws.max_row + 2):
            if deleted_ws.cell(row=row_idx, column=1).value in [None, ""]:
                for col_idx, value in enumerate(source_row_data, start=1):
                    deleted_ws.cell(row=row_idx, column=col_idx).value = value
                break

        wb.save(EXCEL_FILE)
        return jsonify({'success': True})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/get_dropdown_options', methods=['POST'])
def get_dropdown_options():
    """
    Returns dropdown options from the 'Lists' sheet for the frontend
    Expected input JSON: { "sheet": "<sheet_name>" }
    """
    data = request.json
    sheet = data.get('sheet')
    if sheet not in SHEET_NAMES:
        return jsonify({'error': 'Invalid sheet name'}), 400

    try:
        # Read the Lists sheet
        lists_df = pd.read_excel(EXCEL_FILE, sheet_name='Lists')

        # Define expected dropdown columns
        dropdown_columns = [
            "Assessment Status",
            "Well Type",
            "Category",
            "PE/RE",
            "Well Analyst",
            "Current Responsibilities",
            "Servicing Status"
        ]

        dropdown_options = {}

        for col in dropdown_columns:
            if col in lists_df.columns:
                # Drop NaNs, strip whitespace, get unique sorted list
                options = lists_df[col].dropna().astype(str).str.strip().unique()
                dropdown_options[col] = sorted(options.tolist())
            else:
                dropdown_options[col] = []

        return jsonify(dropdown_options)

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/move_to_resolved', methods=['POST'])
def move_to_resolved():
    data = request.json
    sheet_name = data.get('sheet')
    well_name = data.get('well_name')

    if sheet_name not in SHEET_NAMES:
        return jsonify({'error': 'Invalid sheet name'}), 400

    try:
        wb = load_workbook(EXCEL_FILE)
        source_ws = wb[sheet_name]
        resolved_ws = wb['RESOLVED']

        # Get headers from row 2
        headers = [cell.value for cell in source_ws[2]]

        # Find the row with the matching well name
        source_row_data = None
        for row in source_ws.iter_rows(min_row=3, values_only=False):
            cell_value = row[3].value  # Column D (0-indexed as 3)
            if cell_value and str(cell_value).strip() == well_name.strip():
                source_row_data = [cell.value for cell in row]
                source_ws.delete_rows(row[0].row)
                break

        if not source_row_data:
            return jsonify({'error': 'Well not found'}), 404

        # Find the first empty row in Resolved
        for row in range(3, resolved_ws.max_row + 2):
            if resolved_ws.cell(row=row, column=1).value in [None, ""]:
                for col_index, value in enumerate(source_row_data, start=1):
                    resolved_ws.cell(row=row, column=col_index).value = value
                break

        wb.save(EXCEL_FILE)
        return jsonify({'success': True})

    except Exception as e:
        return jsonify({'error': str(e)}), 500
        
        
@app.route('/get_history', methods=['POST'])
def get_history():
    data = request.json
    well_name = data.get('well_name')

    if not well_name:
        return jsonify({'error': 'Well name is required'}), 400

    try:
        wb = load_workbook(EXCEL_FILE)
        resolved_ws = wb['RESOLVED']

        # Get headers from row 2
        headers = [cell.value for cell in resolved_ws[2]]
        try:
            well_col_index = headers.index('Well') + 1
            comments_col_index = headers.index('Comments') + 1  # Column J
        except ValueError:
            return jsonify({'error': 'Well or Comments column not found'}), 400

        history_comments = []

        for row in resolved_ws.iter_rows(min_row=3):
            well_cell = row[well_col_index - 1].value
            comment_cell = row[comments_col_index - 1].value

            if str(well_cell).strip() == well_name.strip() and comment_cell:
                history_comments.append(str(comment_cell).strip())

        return jsonify({'comments': history_comments})

    except Exception as e:
        return jsonify({'error': str(e)}), 500



if __name__ == '__main__':
    app.run(host='0.0.0.0', port=4000, debug=True)
